import openpyxl

wb = openpyxl.load_workbook('簡単仕訳帳2026.xlsx')
print('作成されたシート:', wb.sheetnames)

print('\n=== 月別残高試算表（最初の10行）===')
ws = wb['月別残高試算表']
for i, row in enumerate(ws.iter_rows(max_row=11, values_only=True), 1):
    print(f'Row {i}: {row}')

print('\n=== 総勘定元帳（最初の10行）===')
ws = wb['総勘定元帳']
for i, row in enumerate(ws.iter_rows(max_row=11, values_only=True), 1):
    print(f'Row {i}: {row}')
