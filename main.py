import argparse
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from collections import defaultdict
from datetime import datetime


def load_journal_data(file_path, sheet_name='仕訳帳 ', start_row=5, end_row=309):
    """
    仕訳帳データを読み込む
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 科目コード表を読み込む
    code_sheet = wb['科目コード表']
    code_to_account: dict[int, str] = {}
    for row in code_sheet.iter_rows(min_row=4, max_row=67, values_only=True):  # type: ignore
        code = row[2]  # 列C: 科目コード
        account = row[3]  # 列D: 科目名
        if code is not None and account is not None:
            try:
                code_to_account[int(code)] = str(account)  # type: ignore
            except (ValueError, TypeError):
                pass

    ws = wb[sheet_name]

    data = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=10):  # type: ignore
        values = [cell.value for cell in row]
        # 空行をスキップ（月または日が空の場合）
        if values[1] is None and values[2] is None:
            continue

        # 借方科目と貸方科目をコードから取得
        if values[4] is not None:  # 借方コード
            try:
                values[5] = code_to_account.get(int(values[4]), '')  # type: ignore
            except (ValueError, TypeError):
                values[5] = ''
        if values[7] is not None:  # 貸方コード
            try:
                values[8] = code_to_account.get(int(values[7]), '')  # type: ignore
            except (ValueError, TypeError):
                values[8] = ''

        data.append(values)

    # DataFrameに変換
    columns = ['伝票番号', '月', '日', '摘要', '借方コード', '借方科目', '借方金額',
               '貸方コード', '貸方科目', '貸方金額']
    df = pd.DataFrame(data, columns=columns)

    # データ型を適切に変換
    df['月'] = pd.to_numeric(df['月'], errors='coerce')  # type: ignore
    df['日'] = pd.to_numeric(df['日'], errors='coerce')  # type: ignore
    df['借方金額'] = pd.to_numeric(df['借方金額'], errors='coerce').fillna(0)  # type: ignore
    df['貸方金額'] = pd.to_numeric(df['貸方金額'], errors='coerce').fillna(0)  # type: ignore

    # 空行を削除
    df = df.dropna(subset=['月', '日'])

    return df


def create_trial_balance(df):
    """
    月別残高試算表を作成する
    """
    # 各科目の月別集計（科目コードをキーとして保持）
    trial_balance = defaultdict(lambda: defaultdict(lambda: {'借方': 0, '貸方': 0}))
    account_code_map = {}  # 科目コードと科目名の対応

    for _, row in df.iterrows():
        month = int(row['月'])

        # 借方
        if pd.notna(row['借方科目']) and row['借方金額'] > 0:
            account = row['借方科目']
            code = row['借方コード']
            if pd.notna(code):
                account_code_map[account] = int(code)
            trial_balance[account][month]['借方'] += row['借方金額']

        # 貸方
        if pd.notna(row['貸方科目']) and row['貸方金額'] > 0:
            account = row['貸方科目']
            code = row['貸方コード']
            if pd.notna(code):
                account_code_map[account] = int(code)
            trial_balance[account][month]['貸方'] += row['貸方金額']

    # DataFrameに変換（科目コード順にソート）
    result_data = []
    sorted_accounts = sorted(trial_balance.keys(), key=lambda x: account_code_map.get(x, 9999))
    for account in sorted_accounts:
        for month in range(1, 13):
            if month in trial_balance[account]:
                debit = trial_balance[account][month]['借方']
                credit = trial_balance[account][month]['貸方']
                balance = debit - credit
                code = account_code_map.get(account, '')
                account_display = f"{code} {account}" if code else account
                result_data.append({
                    '月': month,
                    '科目': account_display,
                    '借方合計': debit if debit > 0 else 0,
                    '貸方合計': credit if credit > 0 else 0,
                    '残高': balance
                })

    return pd.DataFrame(result_data)


def create_general_ledger(df):
    """
    総勘定元帳を作成する
    """
    # 科目ごとにデータを整理
    ledger = defaultdict(list)
    account_code_map = {}  # 科目コードと科目名の対応

    for _, row in df.iterrows():
        date_str = f"{int(row['月'])}月{int(row['日'])}日"

        # 借方
        if pd.notna(row['借方科目']) and row['借方金額'] > 0:
            account = row['借方科目']
            code = row['借方コード']
            if pd.notna(code):
                account_code_map[account] = int(code)
            ledger[account].append({
                '日付': date_str,
                '摘要': row['摘要'],
                '相手科目': row['貸方科目'] if pd.notna(row['貸方科目']) else '',
                '借方': row['借方金額'],
                '貸方': 0,
                '月': int(row['月']),
                '日': int(row['日'])
            })

        # 貸方
        if pd.notna(row['貸方科目']) and row['貸方金額'] > 0:
            account = row['貸方科目']
            code = row['貸方コード']
            if pd.notna(code):
                account_code_map[account] = int(code)
            ledger[account].append({
                '日付': date_str,
                '摘要': row['摘要'],
                '相手科目': row['借方科目'] if pd.notna(row['借方科目']) else '',
                '借方': 0,
                '貸方': row['貸方金額'],
                '月': int(row['月']),
                '日': int(row['日'])
            })

    # 各科目のデータを科目コード順、日付順にソートして残高を計算
    ledger_data = []
    sorted_accounts = sorted(ledger.keys(), key=lambda x: account_code_map.get(x, 9999))
    for account in sorted_accounts:
        entries = sorted(ledger[account], key=lambda x: (x['月'], x['日']))
        balance = 0
        code = account_code_map.get(account, '')
        account_display = f"{code} {account}" if code else account
        for entry in entries:
            balance += entry['借方'] - entry['貸方']
            ledger_data.append({
                '科目': account_display,
                '日付': entry['日付'],
                '摘要': entry['摘要'],
                '相手科目': entry['相手科目'],
                '借方': entry['借方'] if entry['借方'] > 0 else None,
                '貸方': entry['貸方'] if entry['貸方'] > 0 else None,
                '残高': balance
            })

    return pd.DataFrame(ledger_data)


def write_to_excel(file_path, trial_balance_df, general_ledger_df):
    """
    既存のExcelファイルに新しいシートを追加
    """
    wb = openpyxl.load_workbook(file_path)

    # 月別残高試算表シートを作成
    if '月別残高試算表' in wb.sheetnames:
        del wb['月別残高試算表']
    ws_trial = wb.create_sheet('月別残高試算表')

    # ヘッダーを設定
    headers = ['月', '科目', '借方合計', '貸方合計', '残高']
    ws_trial.append(headers)

    # データを書き込み
    for row in dataframe_to_rows(trial_balance_df, index=False, header=False):
        ws_trial.append(row)

    # スタイルを適用
    for cell in ws_trial[1]:  # type: ignore
        cell.font = Font(bold=True)  # type: ignore
        cell.alignment = Alignment(horizontal='center')  # type: ignore

    # 総勘定元帳シートを作成
    if '総勘定元帳' in wb.sheetnames:
        del wb['総勘定元帳']
    ws_ledger = wb.create_sheet('総勘定元帳')

    # ヘッダーを設定
    headers = ['科目', '日付', '摘要', '相手科目', '借方', '貸方', '残高']
    ws_ledger.append(headers)

    # データを書き込み
    for row in dataframe_to_rows(general_ledger_df, index=False, header=False):
        ws_ledger.append(row)

    # スタイルを適用
    for cell in ws_ledger[1]:  # type: ignore
        cell.font = Font(bold=True)  # type: ignore
        cell.alignment = Alignment(horizontal='center')  # type: ignore

    # 列幅を調整
    ws_trial.column_dimensions['A'].width = 8  # type: ignore
    ws_trial.column_dimensions['B'].width = 20  # type: ignore
    ws_trial.column_dimensions['C'].width = 15  # type: ignore
    ws_trial.column_dimensions['D'].width = 15  # type: ignore
    ws_trial.column_dimensions['E'].width = 15  # type: ignore

    ws_ledger.column_dimensions['A'].width = 20  # type: ignore
    ws_ledger.column_dimensions['B'].width = 12  # type: ignore
    ws_ledger.column_dimensions['C'].width = 30  # type: ignore
    ws_ledger.column_dimensions['D'].width = 20  # type: ignore
    ws_ledger.column_dimensions['E'].width = 15  # type: ignore
    ws_ledger.column_dimensions['F'].width = 15  # type: ignore
    ws_ledger.column_dimensions['G'].width = 15  # type: ignore

    # ファイルを保存
    wb.save(file_path)
    print(f"処理完了: {file_path}")
    print(f"  - 月別残高試算表: {len(trial_balance_df)} 行")
    print(f"  - 総勘定元帳: {len(general_ledger_df)} 行")


def main():
    parser = argparse.ArgumentParser(
        description='仕訳帳から月別残高試算表と総勘定元帳を生成します'
    )
    parser.add_argument(
        'file',
        nargs='?',
        default='簡単仕訳帳2026.xlsx',
        help='仕訳帳のExcelファイルパス（デフォルト: 簡単仕訳帳2026.xlsx）'
    )
    args = parser.parse_args()
    file_path = args.file

    print(f"仕訳帳データを読み込み中: {file_path}")
    df = load_journal_data(file_path)
    print(f"読み込み完了: {len(df)} 件の仕訳")

    print("\n月別残高試算表を作成中...")
    trial_balance = create_trial_balance(df)

    print("総勘定元帳を作成中...")
    general_ledger = create_general_ledger(df)

    print("\nExcelファイルに書き込み中...")
    write_to_excel(file_path, trial_balance, general_ledger)


if __name__ == "__main__":
    main()
